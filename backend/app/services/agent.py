from __future__ import annotations

import json
from typing import AsyncGenerator

from .llm import get_llm_provider
from .sandbox import execute_code, execute_query
from .excel import build_context

SYSTEM_PROMPT = """你是一个 Excel 数据处理助手。用户会给你 Excel 文件的结构信息和处理需求。

你有两个工具可以使用：

1. **query** - 执行只读 Python 代码来探索数据（查看内容、统计、筛选等）
   参数: {"code": "python代码"}
   代码中可用变量: INPUT_PATH_1, INPUT_PATH_2, ... (每个上传文件一个)
   只需要 print() 输出你想看的内容

2. **execute** - 执行 Python 代码处理数据并生成结果文件
   参数: {"code": "python代码"}
   代码中可用变量: INPUT_PATH_1, INPUT_PATH_2, ..., OUTPUT_PATH
   必须将结果写入 OUTPUT_PATH（支持 .xlsx 格式）

工作流程：
- 先用 query 探索数据，理解内容和结构
- 确认理解后用 execute 生成结果
- 如果执行报错，分析原因后修复代码重试
- 每次只调用一个工具

可用的库：pandas, numpy, re, datetime, openpyxl, copy
禁止使用：os, sys, subprocess, shutil 等系统模块

**极其重要 - 保留原始文件格式：**
当需要修改现有 Excel 文件时，你必须使用 openpyxl 直接操作，**绝对不能**用 pandas 读取后再写回，因为 pandas 会丢失：
- 合并单元格
- 公式
- 样式（字体、颜色、边框、列宽）
- 数据验证
- 隐藏行列
- 注释

正确做法：
```python
import shutil
from openpyxl import load_workbook

# 1. 先复制原始文件到 OUTPUT_PATH，保留所有格式
shutil.copy2(INPUT_PATH_1, OUTPUT_PATH)

# 2. 用 openpyxl 打开副本进行修改
wb = load_workbook(OUTPUT_PATH)
ws = wb['Sheet名称']

# 3. 只修改需要改的单元格
ws['A1'] = '新值'

# 4. 保存
wb.save(OUTPUT_PATH)
```

pandas 只用于**读取和分析**数据（query 工具）。生成结果文件时：
- 修改现有文件 → 用 openpyxl（复制原文件后修改）
- 从零创建新表 → 可以用 pandas

回复时请用中文，简洁说明你在做什么。"""

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "query",
            "description": "执行只读 Python 代码探索数据，返回 print 输出",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "要执行的 Python 代码",
                    }
                },
                "required": ["code"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "execute",
            "description": "执行 Python 代码处理数据并生成结果文件到 OUTPUT_PATH",
            "parameters": {
                "type": "object",
                "properties": {
                    "code": {
                        "type": "string",
                        "description": "要执行的 Python 代码",
                    }
                },
                "required": ["code"],
            },
        },
    },
]

MAX_TURNS = 10
MAX_RETRIES = 3


async def run_agent(
    user_message: str,
    files: list[dict],
) -> AsyncGenerator[dict, None]:
    """
    Agent 主循环，yield SSE 事件：
    - {"type": "text", "content": "..."} 文本流
    - {"type": "tool_call", "name": "query|execute", "code": "..."} 工具调用
    - {"type": "tool_result", "name": "...", "result": "..."} 工具结果
    - {"type": "done", "output_path": "..." | None} 完成
    - {"type": "error", "message": "..."} 错误
    """
    llm = get_llm_provider()

    # 构建文件路径映射
    file_paths = {}
    for i, f in enumerate(files, 1):
        file_paths[f"INPUT_PATH_{i}"] = f["path"]

    # 构建初始消息
    context = build_context(files)
    messages = [
        {"role": "system", "content": SYSTEM_PROMPT},
        {
            "role": "user",
            "content": f"已上传的文件信息：\n\n{context}\n\n用户需求：{user_message}",
        },
    ]

    output_path = None
    retry_count = 0

    for turn in range(MAX_TURNS):
        # 收集完整的 assistant 回复
        full_content = ""
        tool_calls_data: dict[str, dict] = {}
        current_tool_id = None

        async for delta in llm.chat_stream(messages, tools=TOOLS):
            # 文本内容
            if delta.get("content"):
                full_content += delta["content"]
                yield {"type": "text", "content": delta["content"]}

            # 工具调用（流式拼接）
            if "tool_calls" in delta:
                for tc in delta["tool_calls"]:
                    tc_id = tc.get("id")
                    if tc_id:
                        current_tool_id = tc_id
                        tool_calls_data[tc_id] = {
                            "name": tc["function"]["name"],
                            "arguments_str": "",
                        }
                    if current_tool_id and "function" in tc:
                        args_chunk = tc["function"].get("arguments", "")
                        if args_chunk:
                            tool_calls_data[current_tool_id][
                                "arguments_str"
                            ] += args_chunk

        # 没有工具调用 → AI 直接回复，对话结束
        if not tool_calls_data:
            yield {"type": "done", "output_path": output_path}
            return

        # 构建 assistant 消息（含 tool_calls）
        assistant_msg: dict = {
            "role": "assistant",
            "content": full_content or None,
            "tool_calls": [],
        }
        for tc_id, tc_info in tool_calls_data.items():
            assistant_msg["tool_calls"].append(
                {
                    "id": tc_id,
                    "type": "function",
                    "function": {
                        "name": tc_info["name"],
                        "arguments": tc_info["arguments_str"],
                    },
                }
            )
        messages.append(assistant_msg)

        # 逐个执行工具
        for tc_id, tc_info in tool_calls_data.items():
            name = tc_info["name"]
            try:
                args = json.loads(tc_info["arguments_str"])
            except json.JSONDecodeError:
                tool_result = "参数解析失败"
                messages.append(
                    {"role": "tool", "tool_call_id": tc_id, "content": tool_result}
                )
                yield {"type": "tool_result", "name": name, "result": tool_result}
                continue

            code = args.get("code", "")
            yield {"type": "tool_call", "name": name, "code": code}

            if name == "query":
                result = await execute_query(code, file_paths)
                tool_result = (
                    result["output"]
                    if result["success"]
                    else f"错误: {result['output']}"
                )
            elif name == "execute":
                result = await execute_code(code, file_paths)
                if result["success"]:
                    output_path = result["output_path"]
                    tool_result = f"执行成功。输出文件已生成。"
                    if result["stdout"]:
                        tool_result += f"\n标准输出: {result['stdout']}"
                else:
                    retry_count += 1
                    tool_result = (
                        f"执行失败 (重试 {retry_count}/{MAX_RETRIES}): "
                        f"{result['stderr']}"
                    )
                    if retry_count >= MAX_RETRIES:
                        yield {
                            "type": "error",
                            "message": f"代码执行多次失败: {result['stderr']}",
                        }
                        yield {"type": "done", "output_path": None}
                        return
            else:
                tool_result = f"未知工具: {name}"

            messages.append(
                {"role": "tool", "tool_call_id": tc_id, "content": tool_result}
            )
            yield {"type": "tool_result", "name": name, "result": tool_result}

    yield {"type": "error", "message": "达到最大对话轮数"}
    yield {"type": "done", "output_path": output_path}
