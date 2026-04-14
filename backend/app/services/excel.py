from __future__ import annotations

import hashlib
import json
import logging
import os
import uuid
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger("excel-agent.excel")

WORK_DIR = Path(os.environ.get("WORK_DIR", "/data/work"))


def ensure_work_dir() -> None:
    WORK_DIR.mkdir(parents=True, exist_ok=True)


IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".webp")


def save_upload(filename: str, content: bytes) -> dict:
    """保存上传文件，返回文件信息"""
    ensure_work_dir()
    file_id = uuid.uuid4().hex[:12]
    safe_name = f"{file_id}_{filename}"
    file_path = WORK_DIR / safe_name
    file_path.write_bytes(content)

    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if f".{ext}" in IMAGE_EXTS:
        return {
            "file_id": file_id,
            "filename": filename,
            "path": str(file_path),
            "type": "image",
            "profile": None,
        }

    profile = profile_excel(str(file_path))
    return {
        "file_id": file_id,
        "filename": filename,
        "path": str(file_path),
        "type": "excel",
        "profile": profile,
    }


def profile_excel(path: str) -> dict[str, Any]:
    """分析 Excel/CSV 结构：sheet名、列名、类型、样本、行数"""
    if path.endswith(".csv"):
        df = pd.read_csv(path)
        result = {}
        sample_df = df if len(df) <= 200 else df.head(5)
        result["Sheet1"] = {
            "columns": df.columns.tolist(),
            "dtypes": {c: str(df[c].dtype) for c in df.columns},
            "sample": sample_df.fillna("").astype(str).to_dict("records"),
            "row_count": len(df),
            "col_count": len(df.columns),
        }
        return result

    xl = pd.ExcelFile(path)
    result = {}
    for sheet in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        sample_df = df if len(df) <= 200 else df.head(5)
        result[sheet] = {
            "columns": df.columns.tolist(),
            "dtypes": {c: str(df[c].dtype) for c in df.columns},
            "sample": sample_df.fillna("").astype(str).to_dict("records"),
            "row_count": len(df),
            "col_count": len(df.columns),
        }
    return result


def preview_excel(path: str, max_rows: int = 50) -> dict[str, Any]:
    """返回更多行用于前端表格预览"""
    if path.endswith(".csv"):
        df = pd.read_csv(path)
        preview_df = df.head(max_rows)
        return {
            "Sheet1": {
                "columns": df.columns.tolist(),
                "data": preview_df.fillna("").astype(str).values.tolist(),
                "row_count": len(df),
                "col_count": len(df.columns),
            }
        }

    xl = pd.ExcelFile(path)
    result = {}
    for sheet in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        preview_df = df.head(max_rows)
        result[sheet] = {
            "columns": df.columns.tolist(),
            "data": preview_df.fillna("").astype(str).values.tolist(),
            "row_count": len(df),
            "col_count": len(df.columns),
        }
    return result


def compute_diff(input_path: str, output_path: str) -> dict:
    """逐单元格对比输入输出文件，返回结构化 diff。

    使用 openpyxl 读取以正确处理合并单元格和公式。
    行对齐使用第一列作为匹配键。
    """
    wb_in = load_workbook(input_path, data_only=True)
    wb_out = load_workbook(output_path, data_only=True)

    changes = []
    summary = {"modified": 0, "added": 0, "deleted": 0, "unchanged": 0}
    integrity = {"unchanged_rows_ok": True, "type_changes": [], "sum_checks": []}

    # 逐 sheet 对比（以输出文件的 sheet 列表为基准）
    all_sheets = set(wb_in.sheetnames) | set(wb_out.sheetnames)

    for sheet_name in all_sheets:
        if sheet_name not in wb_in.sheetnames:
            # 输出中新增的 sheet
            ws_out = wb_out[sheet_name]
            for row_idx in range(2, ws_out.max_row + 1):  # 跳过表头
                row_data = {}
                for col_idx in range(1, ws_out.max_column + 1):
                    col_name = ws_out.cell(1, col_idx).value or f"列{col_idx}"
                    row_data[str(col_name)] = str(ws_out.cell(row_idx, col_idx).value or "")
                changes.append({"type": "added", "row": row_idx, "sheet": sheet_name, "data": row_data})
                summary["added"] += 1
            continue

        if sheet_name not in wb_out.sheetnames:
            # 输出中删除的 sheet
            ws_in = wb_in[sheet_name]
            for row_idx in range(2, ws_in.max_row + 1):
                row_data = {}
                for col_idx in range(1, ws_in.max_column + 1):
                    col_name = ws_in.cell(1, col_idx).value or f"列{col_idx}"
                    row_data[str(col_name)] = str(ws_in.cell(row_idx, col_idx).value or "")
                changes.append({"type": "deleted", "row": row_idx, "sheet": sheet_name, "data": row_data})
                summary["deleted"] += 1
            continue

        ws_in = wb_in[sheet_name]
        ws_out = wb_out[sheet_name]

        # 读取表头
        max_col = max(ws_in.max_column or 1, ws_out.max_column or 1)
        headers_in = [str(ws_in.cell(1, c).value or f"列{c}") for c in range(1, max_col + 1)]
        headers_out = [str(ws_out.cell(1, c).value or f"列{c}") for c in range(1, max_col + 1)]

        # 读取所有行为 dict（key=第一列值）
        def read_rows(ws, headers, max_c):
            rows = {}
            row_order = []
            for r in range(2, (ws.max_row or 1) + 1):
                vals = [ws.cell(r, c).value for c in range(1, max_c + 1)]
                if all(v is None for v in vals):
                    continue
                key = str(vals[0]) if vals[0] is not None else f"__row_{r}"
                # 处理重复 key
                if key in rows:
                    key = f"{key}__dup_{r}"
                row_dict = {headers[i]: vals[i] for i in range(min(len(headers), len(vals)))}
                rows[key] = {"row_num": r, "data": row_dict}
                row_order.append(key)
            return rows, row_order

        in_rows, in_order = read_rows(ws_in, headers_in, max_col)
        out_rows, out_order = read_rows(ws_out, headers_out, max_col)

        in_keys = set(in_order)
        out_keys = set(out_order)

        # 第一列作为 row_label
        first_col = headers_in[0] if headers_in else "列1"

        # deleted 行
        for key in in_order:
            if key not in out_keys:
                r = in_rows[key]
                label = str(r["data"].get(first_col, key))
                changes.append({
                    "type": "deleted",
                    "row": r["row_num"],
                    "sheet": sheet_name,
                    "row_label": f"「{label}」(第{r['row_num']}行)",
                    "data": {k: str(v or "") for k, v in r["data"].items()},
                })
                summary["deleted"] += 1

        # added 行
        for key in out_order:
            if key not in in_keys:
                r = out_rows[key]
                changes.append({
                    "type": "added",
                    "row": r["row_num"],
                    "sheet": sheet_name,
                    "data": {k: str(v or "") for k, v in r["data"].items()},
                })
                summary["added"] += 1

        # matched 行：逐单元格对比
        for key in in_order:
            if key not in out_keys:
                continue
            in_row = in_rows[key]
            out_row = out_rows[key]
            row_modified = False
            all_cols = set(list(in_row["data"].keys()) + list(out_row["data"].keys()))

            for col in all_cols:
                old_val = in_row["data"].get(col)
                new_val = out_row["data"].get(col)
                if str(old_val or "") != str(new_val or ""):
                    row_modified = True
                    label = str(in_row["data"].get(first_col, key))
                    changes.append({
                        "type": "modified",
                        "row": in_row["row_num"],
                        "sheet": sheet_name,
                        "col_name": col,
                        "row_label": f"「{label}」(第{in_row['row_num']}行)",
                        "old": str(old_val or ""),
                        "new": str(new_val or ""),
                    })
                    summary["modified"] += 1

            if not row_modified:
                summary["unchanged"] += 1
            else:
                # 验证"未修改"的单元格
                pass

        # 未修改行 hash 校验
        unchanged_ok = True
        for key in in_order:
            if key not in out_keys:
                continue
            in_row = in_rows[key]
            out_row = out_rows[key]
            # 检查是否有任何单元格变化（已在上面统计）
            in_hash = hashlib.md5(
                json.dumps({k: str(v or "") for k, v in in_row["data"].items()}, sort_keys=True).encode()
            ).hexdigest()
            out_hash = hashlib.md5(
                json.dumps({k: str(v or "") for k, v in out_row["data"].items()}, sort_keys=True).encode()
            ).hexdigest()
            # 如果被标记为 unchanged 但 hash 不同，说明有误改
            all_cols = set(list(in_row["data"].keys()) + list(out_row["data"].keys()))
            has_change = any(
                str(in_row["data"].get(c) or "") != str(out_row["data"].get(c) or "")
                for c in all_cols
            )
            if not has_change and in_hash != out_hash:
                unchanged_ok = False

        integrity["unchanged_rows_ok"] = unchanged_ok

        # 数值列 sum 校验（用 pandas）
        try:
            df_in = pd.read_excel(input_path, sheet_name=sheet_name)
            df_out = pd.read_excel(output_path, sheet_name=sheet_name)
            for col in df_in.select_dtypes(include=["number"]).columns:
                if col in df_out.columns:
                    before = float(df_in[col].sum())
                    after = float(df_out[col].sum())
                    if abs(before - after) > 0.001:
                        diff_val = after - before
                        integrity["sum_checks"].append({
                            "column": col,
                            "before": round(before, 2),
                            "after": round(after, 2),
                            "diff": round(diff_val, 2),
                        })
        except Exception as e:
            logger.warning("sum校验异常: %s", e)

    wb_in.close()
    wb_out.close()

    # 超长改动截断
    truncated = False
    total_changes = len(changes)
    if total_changes > 50:
        changes = changes[:50]
        truncated = True

    return {
        "summary": summary,
        "integrity": integrity,
        "changes": changes,
        "truncated": truncated,
        "total_changes": total_changes,
    }


def compute_create_summary(output_path: str) -> dict:
    """生成新建文件的摘要信息"""
    try:
        if output_path.endswith(".csv"):
            df = pd.read_csv(output_path)
            return {
                "sheets": {
                    "Sheet1": {
                        "row_count": len(df),
                        "col_count": len(df.columns),
                        "columns": df.columns.tolist(),
                        "preview": df.head(5).fillna("").astype(str).to_dict("records"),
                    }
                }
            }

        xl = pd.ExcelFile(output_path)
        sheets = {}
        for sheet in xl.sheet_names:
            df = pd.read_excel(output_path, sheet_name=sheet)
            sheets[sheet] = {
                "row_count": len(df),
                "col_count": len(df.columns),
                "columns": df.columns.tolist(),
                "preview": df.head(5).fillna("").astype(str).to_dict("records"),
            }
        return {"sheets": sheets}
    except Exception as e:
        logger.error("create摘要生成失败: %s", e)
        return {"error": str(e)}


def generate_operation_summary(user_message: str, diff: dict) -> str:
    """从 diff 数据生成一句话操作摘要（模板拼接，不靠 LLM）"""
    parts = [user_message[:50]]
    s = diff.get("summary", {})
    details = []
    if s.get("modified"):
        details.append(f"改了{s['modified']}个格子")
    if s.get("added"):
        details.append(f"新增{s['added']}行")
    if s.get("deleted"):
        details.append(f"删除{s['deleted']}行")
    if details:
        parts.append(f"（{'，'.join(details)}）")
    return "".join(parts)


def build_context(files: list[dict]) -> str:
    """根据文件 profile 构建给 LLM 的上下文描述"""
    parts = []
    idx = 0
    for f in files:
        if f.get("type") == "image" or not f.get("profile"):
            continue
        idx += 1
        parts.append(f"## 文件 {idx}: {f['filename']} (变量名: INPUT_PATH_{idx})")
        for sheet_name, info in f["profile"].items():
            total_rows = info["row_count"]
            parts.append(f"\n### Sheet: {sheet_name} ({total_rows} 行, {info['col_count']} 列)")
            parts.append(f"列: {info['columns']}")
            parts.append(f"类型: {info['dtypes']}")
            if total_rows <= 200:
                parts.append(f"数据(全量):\n{info['sample']}")
            else:
                parts.append(f"样本(前5行):\n{info['sample']}")
                parts.append("(可用 query 工具探索更多数据)")
    return "\n".join(parts)
